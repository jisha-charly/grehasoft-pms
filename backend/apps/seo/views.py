import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone
from django.db import transaction
from django.db.models import Sum, Q, Count, Avg
from django.http import HttpResponse
from django.contrib.auth import get_user_model

from rest_framework import viewsets, status, permissions
from rest_framework.pagination import PageNumberPagination
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied, ValidationError

from .models import (
    SEOActivityType, Website, Keyword, SEODailyWorkLog, SEODailyWorkLogItem, SEOMonthlyTarget, SEOTask, SEOReminder, SEOCredential, SEODailyWorkProof
)
from .serializers import (
    SEOActivityTypeSerializer, WebsiteSerializer, KeywordSerializer,
    SEODailyWorkLogSerializer, SEOMonthlyTargetSerializer, SEOTaskSerializer, SEOReminderSerializer, SEOCredentialSerializer
)
from .utils import decrypt_password, encrypt_password
from apps.projects.models import Client
from apps.projects.utils import log_failed_attempt



def is_admin(user):
    return user.is_authenticated and (user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == "SUPER_ADMIN"))


def is_seo_manager(user):
    return user.is_authenticated and hasattr(user, 'role') and user.role and user.role.name == "SEO_MANAGER"


def is_seo_executive(user):
    return user.is_authenticated and hasattr(user, 'role') and user.role and user.role.name == "SEO_EXECUTIVE"


class SEOActivityTypeViewSet(viewsets.ModelViewSet):
    queryset = SEOActivityType.objects.all()
    serializer_class = SEOActivityTypeSerializer

    def get_queryset(self):
        user = self.request.user
        qs = SEOActivityType.objects.all()

        is_privileged = is_admin(user) or is_seo_manager(user)
        if not is_privileged:
            qs = qs.filter(is_active=True)
        else:
            active_param = self.request.query_params.get("active")
            if active_param is not None:
                is_active_val = active_param.lower() in ["true", "1", "yes"]
                qs = qs.filter(is_active=is_active_val)

        return qs

    def get_permissions(self):
        return [permissions.IsAuthenticated()]

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.method not in permissions.SAFE_METHODS:
            user = request.user
            if not (is_admin(user) or is_seo_manager(user)):
                raise PermissionDenied("You do not have permission to manage SEO activity types.")


class WebsiteViewSet(viewsets.ModelViewSet):
    queryset = Website.objects.all()
    serializer_class = WebsiteSerializer

    def get_queryset(self):
        user = self.request.user
        role_name = getattr(user.role, 'name', None) if hasattr(user, 'role') else None
        if is_admin(user) or is_seo_manager(user):
            return Website.objects.all().select_related("client", "assigned_executive", "assigned_by")
        elif is_seo_executive(user):
            return Website.objects.filter(assigned_executive=user).select_related("client", "assigned_executive", "assigned_by")
        elif role_name == 'CLIENT':
            client = user.get_associated_client()
            if client:
                return Website.objects.filter(client=client).select_related("client", "assigned_executive", "assigned_by")
        return Website.objects.none()

    def check_permissions(self, request):
        super().check_permissions(request)
        role_name = getattr(request.user.role, 'name', None) if hasattr(request.user, 'role') else None
        if role_name == 'CLIENT' and request.method not in permissions.SAFE_METHODS:
            log_failed_attempt(request.user, f"Tried to write website via {request.method}")
            self.permission_denied(request, message="Clients do not have permission to modify websites.")

    def perform_create(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can create websites.")
        
        assigned_executive = serializer.validated_data.get("assigned_executive")
        if assigned_executive:
            serializer.save(assigned_by=user, assigned_date=timezone.now())
        else:
            serializer.save()

    def perform_update(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can update websites.")

        instance = self.get_object()
        assigned_executive = serializer.validated_data.get("assigned_executive")
        
        if assigned_executive and instance.assigned_executive != assigned_executive:
            serializer.save(assigned_by=user, assigned_date=timezone.now())
        else:
            serializer.save()


class KeywordViewSet(viewsets.ModelViewSet):
    queryset = Keyword.objects.all()
    serializer_class = KeywordSerializer

    def get_queryset(self):
        user = self.request.user
        website_id = self.request.query_params.get("website")
        role_name = getattr(user.role, 'name', None) if hasattr(user, 'role') else None
        
        # Base filter based on user roles
        if is_admin(user) or is_seo_manager(user):
            qs = Keyword.objects.all().select_related("website")
        elif is_seo_executive(user):
            qs = Keyword.objects.filter(website__assigned_executive=user).select_related("website")
        elif role_name == 'CLIENT':
            client = user.get_associated_client()
            if client:
                qs = Keyword.objects.filter(website__client=client).select_related("website")
            else:
                qs = Keyword.objects.none()
        else:
            qs = Keyword.objects.none()

        if website_id:
            qs = qs.filter(website_id=website_id)
        return qs

    def check_permissions(self, request):
        super().check_permissions(request)
        role_name = getattr(request.user.role, 'name', None) if hasattr(request.user, 'role') else None
        if role_name == 'CLIENT' and request.method not in permissions.SAFE_METHODS:
            log_failed_attempt(request.user, f"Tried to write keyword via {request.method}")
            self.permission_denied(request, message="Clients do not have permission to modify target keywords.")

    def perform_create(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can add target keywords.")
        serializer.save()


class SEODailyWorkLogViewSet(viewsets.ModelViewSet):
    queryset = SEODailyWorkLog.objects.all()
    serializer_class = SEODailyWorkLogSerializer

    def get_queryset(self):
        user = self.request.user
        role_name = getattr(user.role, 'name', None) if hasattr(user, 'role') else None
        if is_admin(user) or is_seo_manager(user):
            qs = SEODailyWorkLog.objects.all().select_related(
                "website", "executive", "created_by", "updated_by", "approved_by", "rejected_by"
            ).prefetch_related("items__activity_type")
        elif is_seo_executive(user):
            qs = SEODailyWorkLog.objects.filter(executive=user).select_related(
                "website", "executive", "created_by", "updated_by", "approved_by", "rejected_by"
            ).prefetch_related("items__activity_type")
        elif role_name == 'CLIENT':
            client = user.get_associated_client()
            if client:
                qs = SEODailyWorkLog.objects.filter(website__client=client, status="approved").select_related(
                    "website", "executive", "created_by", "updated_by", "approved_by", "rejected_by"
                ).prefetch_related("items__activity_type")
            else:
                qs = SEODailyWorkLog.objects.none()
        else:
            qs = SEODailyWorkLog.objects.none()

        # Filtering
        status_param = self.request.query_params.get("status")
        website_id = self.request.query_params.get("website")
        executive_id = self.request.query_params.get("executive")
        activity_type_id = self.request.query_params.get("activity_type")
        keyword = self.request.query_params.get("keyword")
        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")

        if status_param:
            qs = qs.filter(status=status_param)
        if website_id:
            qs = qs.filter(website_id=website_id)
        if executive_id:
            qs = qs.filter(executive_id=executive_id)
        if activity_type_id:
            qs = qs.filter(items__activity_type_id=activity_type_id)
        if keyword:
            qs = qs.filter(items__keyword__icontains=keyword)
        if start_date:
            qs = qs.filter(log_date__gte=start_date)
        if end_date:
            qs = qs.filter(log_date__lte=end_date)

        return qs.distinct()

    def check_permissions(self, request):
        super().check_permissions(request)
        role_name = getattr(request.user.role, 'name', None) if hasattr(request.user, 'role') else None
        if role_name == 'CLIENT' and request.method not in permissions.SAFE_METHODS:
            log_failed_attempt(request.user, f"Tried to write daily work log via {request.method}")
            self.permission_denied(request, message="Clients do not have permission to modify daily work logs.")

    def create(self, request, *args, **kwargs):
        user = request.user
        website_id = request.data.get("website")
        log_date = request.data.get("log_date")

        if website_id and log_date:
            existing_log = SEODailyWorkLog.objects.filter(
                executive=user,
                website_id=website_id,
                log_date=log_date
            ).first()

            if existing_log:
                items_data = request.data.get("items", [])
                if isinstance(items_data, str):
                    import json
                    try:
                        items_data = json.loads(items_data)
                    except ValueError:
                        pass

                for item_data in items_data:
                    password = item_data.get("password")
                    encrypted_password = None
                    if password:
                        encrypted_password = encrypt_password(password)

                    SEODailyWorkLogItem.objects.create(
                        work_log=existing_log,
                        activity_type_id=item_data.get("activity_type"),
                        count=item_data.get("count", 1),
                        keyword=item_data.get("keyword"),
                        submission_url=item_data.get("submission_url"),
                        domain_authority=item_data.get("domain_authority"),
                        spam_score=item_data.get("spam_score"),
                        time_spent_minutes=item_data.get("time_spent_minutes"),
                        username=item_data.get("username"),
                        password=encrypted_password
                    )

                remarks = request.data.get("remarks")
                if remarks:
                    existing_log.remarks = remarks

                status_param = request.data.get("status")
                if status_param:
                    existing_log.status = status_param

                proof_file = request.FILES.get("proof_file")
                if proof_file:
                    max_size = 10 * 1024 * 1024
                    if proof_file.size > max_size:
                        raise ValidationError({"proof_file": "File size exceeds 10MB limit."})
                    ext = proof_file.name.split('.')[-1].lower()
                    allowed = ['xls', 'xlsx', 'pdf', 'jpg', 'jpeg', 'png', 'zip']
                    if ext not in allowed:
                        raise ValidationError({"proof_file": "File type not supported. Allowed: xls, xlsx, pdf, jpg, jpeg, png, zip"})

                    existing_log.proof_file = proof_file
                    SEODailyWorkProof.objects.create(work_log=existing_log, proof_file=proof_file)

                existing_log.calculate_total_count()
                existing_log.save()

                serializer = self.get_serializer(existing_log)
                return Response(serializer.data, status=status.HTTP_201_CREATED)

        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        log = serializer.save(executive=self.request.user, created_by=self.request.user)
        if log.proof_file:
            SEODailyWorkProof.objects.create(work_log=log, proof_file=log.proof_file)

    def perform_update(self, serializer):
        instance = self.get_object()
        user = self.request.user

        if is_seo_executive(user):
            if instance.executive != user:
                raise PermissionDenied("You can only edit your own logs.")
            if instance.status not in ["draft", "rejected"]:
                if "proof_file" in self.request.FILES:
                    pass
                else:
                    raise PermissionDenied("You can only edit logs in Draft or Rejected status.")

            status_param = self.request.data.get("status")
            if status_param and status_param not in ["draft", "submitted"]:
                raise ValidationError({"status": "You can only save as draft or submit."})

        log = serializer.save(updated_by=user)
        proof_file = self.request.FILES.get("proof_file")
        if proof_file:
            SEODailyWorkProof.objects.create(work_log=log, proof_file=proof_file)

    @action(detail=False, methods=["post"], url_path="add-items")
    def add_items(self, request):
        user = request.user
        website_id = request.data.get("website")
        log_date = request.data.get("log_date")

        if not website_id or not log_date:
            return Response({"error": "website and log_date are required."}, status=status.HTTP_400_BAD_REQUEST)

        log, created = SEODailyWorkLog.objects.get_or_create(
            executive=user,
            website_id=website_id,
            log_date=log_date,
            defaults={
                "status": request.data.get("status", "submitted"),
                "remarks": request.data.get("remarks", ""),
                "created_by": user
            }
        )

        if not created:
            remarks = request.data.get("remarks")
            if remarks:
                log.remarks = remarks
            status_param = request.data.get("status")
            if status_param:
                log.status = status_param

        proof_file = request.FILES.get("proof_file")
        if proof_file:
            max_size = 10 * 1024 * 1024
            if proof_file.size > max_size:
                raise ValidationError({"proof_file": "File size exceeds 10MB limit."})
            ext = proof_file.name.split('.')[-1].lower()
            allowed = ['xls', 'xlsx', 'pdf', 'jpg', 'jpeg', 'png', 'zip']
            if ext not in allowed:
                raise ValidationError({"proof_file": "File type not supported. Allowed: xls, xlsx, pdf, jpg, jpeg, png, zip"})

            log.proof_file = proof_file
            SEODailyWorkProof.objects.create(work_log=log, proof_file=proof_file)

        items_data = request.data.get("items", [])
        if isinstance(items_data, str):
            import json
            try:
                items_data = json.loads(items_data)
            except ValueError:
                pass

        for item_data in items_data:
            password = item_data.get("password")
            encrypted_password = None
            if password:
                encrypted_password = encrypt_password(password)

            SEODailyWorkLogItem.objects.create(
                work_log=log,
                activity_type_id=item_data.get("activity_type"),
                count=item_data.get("count", 1),
                keyword=item_data.get("keyword"),
                submission_url=item_data.get("submission_url"),
                domain_authority=item_data.get("domain_authority"),
                spam_score=item_data.get("spam_score"),
                time_spent_minutes=item_data.get("time_spent_minutes"),
                username=item_data.get("username"),
                password=encrypted_password
            )

        log.calculate_total_count()
        log.save()

        serializer = self.get_serializer(log)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        user = request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can approve logs.")
            
        log = self.get_object()
        log.status = "approved"
        log.approved_by = user
        log.approved_date = timezone.now()
        log.save()
        return Response({"status": "approved"})

    @action(detail=True, methods=["post"])
    def reject(self, request, pk=None):
        user = request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can reject logs.")
            
        remarks = request.data.get("remarks_by_manager")
        if not remarks:
            return Response({"error": "Rejection remarks are required."}, status=400)
            
        log = self.get_object()
        log.status = "rejected"
        log.remarks_by_manager = remarks
        log.rejected_by = user
        log.rejected_date = timezone.now()
        log.save()
        return Response({"status": "rejected"})

    @action(detail=False, methods=["get"])
    def dashboard(self, request):
        user = request.user
        today = datetime.date.today()
        current_month = today.strftime("%Y-%m")

        if is_admin(user) or is_seo_manager(user):
            # Manager Dashboard KPIs
            total_clients = Client.objects.filter(seo_websites__isnull=False).distinct().count()
            total_websites = Website.objects.count()
            active_projects = Website.objects.filter(status="active").count()
            
            logs_today = SEODailyWorkLog.objects.filter(log_date=today)
            activities_today = SEODailyWorkLogItem.objects.filter(work_log__log_date=today).aggregate(Sum('count'))['count__sum'] or 0
            
            # Fetch for current month
            logs_this_month = SEODailyWorkLog.objects.filter(log_date__year=today.year, log_date__month=today.month)
            activities_this_month = SEODailyWorkLogItem.objects.filter(
                work_log__log_date__year=today.year, 
                work_log__log_date__month=today.month
            ).aggregate(Sum('count'))['count__sum'] or 0
            
            pending_reviews = SEODailyWorkLog.objects.filter(status="submitted").count()

            # Targets and completion progress
            targets = SEOMonthlyTarget.objects.filter(month=current_month)
            total_target = targets.aggregate(Sum('target_count'))['target_count__sum'] or 0
            
            # Target vs Completed (Approved/Submitted counts)
            completed_logs = SEODailyWorkLog.objects.filter(
                log_date__year=today.year, 
                log_date__month=today.month, 
                status__in=["submitted", "approved"]
            )
            completed_count = SEODailyWorkLogItem.objects.filter(work_log__in=completed_logs).aggregate(Sum('count'))['count__sum'] or 0

            target_pct = round((completed_count / total_target * 100), 2) if total_target > 0 else 0

            # Top Performing Executive
            top_exec = SEODailyWorkLogItem.objects.filter(
                work_log__log_date__year=today.year, 
                work_log__log_date__month=today.month
            ).values(
                'work_log__executive__name', 'work_log__executive__username'
            ).annotate(
                total_work=Sum('count')
            ).order_by('-total_work').first()

            top_exec_name = "N/A"
            if top_exec:
                top_exec_name = top_exec['work_log__executive__name'] or top_exec['work_log__executive__username']

            # Monthly Trend
            monthly_trend = []
            for m in range(1, 13):
                date_filter = datetime.date(today.year, m, 1)
                month_str = date_filter.strftime("%b")
                month_cnt = SEODailyWorkLogItem.objects.filter(
                    work_log__log_date__year=today.year, 
                    work_log__log_date__month=m
                ).aggregate(Sum('count'))['count__sum'] or 0
                monthly_trend.append({"month": month_str, "activities": month_cnt})

            # Activities by Type
            by_type = SEODailyWorkLogItem.objects.filter(
                work_log__log_date__year=today.year, 
                work_log__log_date__month=today.month
            ).values('activity_type__name').annotate(count=Sum('count')).order_by('-count')

            by_website = SEODailyWorkLogItem.objects.filter(
                work_log__log_date__year=today.year, 
                work_log__log_date__month=today.month
            ).values('work_log__website__website_name').annotate(count=Sum('count')).order_by('-count')[:5]

            by_exec = SEODailyWorkLogItem.objects.filter(
                work_log__log_date__year=today.year, 
                work_log__log_date__month=today.month
            ).values('work_log__executive__name', 'work_log__executive__username').annotate(count=Sum('count')).order_by('-count')

            return Response({
                "kpis": {
                    "total_clients": total_clients,
                    "total_websites": total_websites,
                    "active_projects": active_projects,
                    "activities_today": activities_today,
                    "activities_this_month": activities_this_month,
                    "pending_reviews": pending_reviews,
                    "target_completion_pct": target_pct,
                    "top_performing_executive": top_exec_name
                },
                "monthly_trend": monthly_trend,
                "activities_by_type": [{"type": x['activity_type__name'], "count": x['count']} for x in by_type],
                "activities_by_website": [{"website": x['work_log__website__website_name'], "count": x['count']} for x in by_website],
                "activities_by_executive": [{"executive": x['work_log__executive__name'] or x['work_log__executive__username'], "count": x['count']} for x in by_exec]
            })

        else:
            # Executive Dashboard KPIs
            my_logs_today = SEODailyWorkLog.objects.filter(executive=user, log_date=today)
            today_count = SEODailyWorkLogItem.objects.filter(work_log__in=my_logs_today).aggregate(Sum('count'))['count__sum'] or 0

            # Targets progress
            my_targets = SEOMonthlyTarget.objects.filter(executive=user, month=current_month)
            my_target_sum = my_targets.aggregate(Sum('target_count'))['target_count__sum'] or 0
            
            my_logs_month = SEODailyWorkLog.objects.filter(
                executive=user, 
                log_date__year=today.year, 
                log_date__month=today.month,
                status__in=["submitted", "approved"]
            )
            my_month_count = SEODailyWorkLogItem.objects.filter(work_log__in=my_logs_month).aggregate(Sum('count'))['count__sum'] or 0

            progress_pct = round((my_month_count / my_target_sum * 100), 2) if my_target_sum > 0 else 0

            assigned_websites_cnt = Website.objects.filter(assigned_executive=user, status="active").count()
            pending_tasks_cnt = SEOTask.objects.filter(assigned_executive=user, status="pending").count()
            pending_reminders_cnt = SEOReminder.objects.filter(assigned_executive=user, status="pending").count()

            # Targets list
            targets_list = []
            for tg in my_targets:
                completed = SEODailyWorkLogItem.objects.filter(
                    work_log__executive=user,
                    work_log__log_date__year=today.year,
                    work_log__log_date__month=today.month,
                    work_log__status__in=["submitted", "approved"],
                    activity_type=tg.activity_type
                ).aggregate(Sum('count'))['count__sum'] or 0
                targets_list.append({
                    "activity_type": tg.activity_type.name,
                    "target": tg.target_count,
                    "completed": completed,
                    "progress": round((completed / tg.target_count * 100), 2) if tg.target_count > 0 else 0
                })

            return Response({
                "kpis": {
                    "today_count": today_count,
                    "progress_pct": progress_pct,
                    "assigned_websites": assigned_websites_cnt,
                    "pending_tasks": pending_tasks_cnt,
                    "pending_reminders": pending_reminders_cnt
                },
                "targets": targets_list
            })

    @action(detail=False, methods=["get"])
    def team_performance(self, request):
        user = request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can view team performance.")

        today = datetime.date.today()
        start_of_week = today - datetime.timedelta(days=today.weekday())
        current_month = today.strftime("%Y-%m")
        
        User = get_user_model()
        # Query SEO Executives (users whose role is SEO_EXECUTIVE)
        executives = User.objects.filter(role__name="SEO_EXECUTIVE")
        
        leaderboard = []
        for exec_user in executives:
            # Activities logged today
            today_logs = SEODailyWorkLog.objects.filter(executive=exec_user, log_date=today)
            cnt_today = SEODailyWorkLogItem.objects.filter(work_log__in=today_logs).aggregate(Sum('count'))['count__sum'] or 0

            # Weekly activities
            week_logs = SEODailyWorkLog.objects.filter(executive=exec_user, log_date__gte=start_of_week)
            cnt_week = SEODailyWorkLogItem.objects.filter(work_log__in=week_logs).aggregate(Sum('count'))['count__sum'] or 0

            # Monthly activities
            month_logs = SEODailyWorkLog.objects.filter(executive=exec_user, log_date__year=today.year, log_date__month=today.month)
            cnt_month = SEODailyWorkLogItem.objects.filter(work_log__in=month_logs).aggregate(Sum('count'))['count__sum'] or 0

            # Approval Rate
            reviewed_logs = month_logs.filter(status__in=["approved", "rejected"])
            reviewed_cnt = SEODailyWorkLogItem.objects.filter(work_log__in=reviewed_logs).aggregate(Sum('count'))['count__sum'] or 0
            
            approved_logs = month_logs.filter(status="approved")
            approved_cnt = SEODailyWorkLogItem.objects.filter(work_log__in=approved_logs).aggregate(Sum('count'))['count__sum'] or 0

            approval_rate = 100
            if reviewed_cnt > 0:
                approval_rate = round((approved_cnt / reviewed_cnt * 100), 2)

            assigned_sites = Website.objects.filter(assigned_executive=exec_user, status="active").count()

            # Breakdown by activity type
            breakdown_query = SEODailyWorkLogItem.objects.filter(
                work_log__executive=exec_user,
                work_log__log_date__year=today.year,
                work_log__log_date__month=today.month
            ).values('activity_type__name').annotate(total=Sum('count'))
            
            breakdown = {x['activity_type__name']: x['total'] for x in breakdown_query}

            leaderboard.append({
                "id": exec_user.id,
                "name": exec_user.name or exec_user.username,
                "activities_today": cnt_today,
                "activities_this_week": cnt_week,
                "activities_this_month": cnt_month,
                "approval_rate": approval_rate,
                "assigned_websites": assigned_sites,
                "breakdown": breakdown
            })

        # Sort leaderboard by monthly activity count descending
        leaderboard.sort(key=lambda x: x["activities_this_month"], reverse=True)
        return Response(leaderboard)

    @action(detail=False, methods=["post"])
    def import_excel(self, request):
        if not (is_admin(request.user) or is_seo_manager(request.user)):
            raise PermissionDenied("Only SEO Managers and Admins can import work logs.")
            
        file_obj = request.FILES.get("file")
        if not file_obj:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response({"error": f"Invalid Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_row:
            return Response({"error": "Empty Excel sheet."}, status=status.HTTP_400_BAD_REQUEST)

        # Create a mapping of lowercase header titles to their column index
        header_map = {str(val).strip().lower(): idx for idx, val in enumerate(header_row) if val is not None}

        # Flexible mapping dictionary
        flex_map = {
            "website": ["website"],
            "date": ["date"],
            "activity_type": ["type of submission", "activity type", "submission type", "activity"],
            "keyword": ["keyword", "target keyword"],
            "submission_url": ["submission url", "url", "live url", "link"],
            "domain_authority": ["domain", "domain authority", "da"],
            "spam_score": ["spam", "spam score", "ss"],
            "time_spent_minutes": ["time", "time spent", "duration", "time spent (minutes)"],
            "count": ["count", "quantity"],
            "executive_username": ["executive username", "executive", "username"]
        }

        # Resolve header mapping to flex keys
        resolved_headers = {}
        for key, aliases in flex_map.items():
            for alias in aliases:
                if alias in header_map:
                    resolved_headers[key] = header_map[alias]
                    break

        # Check required fields
        required_keys = ["website", "date", "activity_type"]
        missing = [k for k in required_keys if k not in resolved_headers]
        if missing:
            return Response({"error": f"Missing required columns (or aliases): {', '.join(missing)}. Please check your headers."}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        logs_to_create = []
        User = get_user_model()

        row_num = 1
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not any(row_vals):
                continue

            def get_val_flex(key, default=None):
                idx = resolved_headers.get(key)
                return row_vals[idx] if idx is not None else default

            web_val = get_val_flex("website")
            date_val = get_val_flex("date")
            act_type_val = get_val_flex("activity_type")
            
            kw_val = get_val_flex("keyword", "")
            sub_url_val = get_val_flex("submission_url", "")
            da_val = get_val_flex("domain_authority")
            spam_val = get_val_flex("spam_score")
            time_val = get_val_flex("time_spent_minutes")
            
            # Count default is auto-calculated from URLs if bulk urls are pasted.
            url_str = str(sub_url_val or "").strip()
            urls_list = [u.strip() for u in url_str.split("\n") if u.strip()]
            url_count = len(urls_list) if urls_list else 1
            
            count_val = get_val_flex("count", url_count)
            exec_username = get_val_flex("executive_username")

            if not web_val or not date_val or not act_type_val:
                errors.append(f"Row {row_num}: Website, Date, and Activity Type are required.")
                continue

            website = Website.objects.filter(
                Q(website_name__iexact=str(web_val).strip()) | Q(domain_url__icontains=str(web_val).strip())
            ).first()
            if not website:
                errors.append(f"Row {row_num}: Website '{web_val}' not found in database.")
                continue

            activity_type = SEOActivityType.objects.filter(name__iexact=str(act_type_val).strip()).first()
            if not activity_type:
                errors.append(f"Row {row_num}: Activity Type '{act_type_val}' is not registered.")
                continue

            parsed_date = None
            if isinstance(date_val, datetime.date):
                parsed_date = date_val
            elif isinstance(date_val, datetime.datetime):
                parsed_date = date_val.date()
            else:
                try:
                    parsed_date = datetime.datetime.strptime(str(date_val).strip(), "%Y-%m-%d").date()
                except ValueError:
                    try:
                        parsed_date = datetime.datetime.strptime(str(date_val).strip(), "%d-%m-%Y").date()
                    except ValueError:
                        errors.append(f"Row {row_num}: Invalid date format '{date_val}'. Use YYYY-MM-DD.")
                        continue

            executive = None
            if exec_username:
                executive = User.objects.filter(username=str(exec_username).strip()).first()
                if not executive:
                    errors.append(f"Row {row_num}: Executive user '{exec_username}' not found.")
                    continue
            else:
                executive = website.assigned_executive or request.user

            def parse_int_optional(val, name):
                if val is None or str(val).strip() == "":
                    return None
                try:
                    return int(float(val))
                except ValueError:
                    errors.append(f"Row {row_num}: {name} must be an integer number.")
                    return "error"

            da = parse_int_optional(da_val, "Domain Authority (DA)")
            spam = parse_int_optional(spam_val, "Spam Score")
            time_spent = parse_int_optional(time_val, "Time Spent")
            count = parse_int_optional(count_val, "Count")

            if "error" in [da, spam, time_spent, count]:
                continue

            logs_to_create.append({
                "website": website,
                "log_date": parsed_date,
                "executive": executive,
                "activity_type": activity_type,
                "count": count if count is not None else 1,
                "keyword": kw_val,
                "submission_url": sub_url_val,
                "domain_authority": da,
                "spam_score": spam,
                "time_spent_minutes": time_spent,
                "remarks": f"Imported work log."
            })

        if errors:
            return Response({"errors": errors}, status=status.HTTP_400_BAD_REQUEST)

        imported_count = 0
        try:
            with transaction.atomic():
                for item in logs_to_create:
                    log, created = SEODailyWorkLog.objects.get_or_create(
                        executive=item["executive"],
                        website=item["website"],
                        log_date=item["log_date"],
                        defaults={
                            "remarks": item["remarks"],
                            "status": "submitted",
                            "created_by": request.user
                        }
                    )
                    SEODailyWorkLogItem.objects.create(
                        work_log=log,
                        activity_type=item["activity_type"],
                        count=item["count"],
                        keyword=item["keyword"],
                        submission_url=item["submission_url"],
                        domain_authority=item["domain_authority"],
                        spam_score=item["spam_score"],
                        time_spent_minutes=item["time_spent_minutes"]
                    )
                    log.calculate_total_count()
                    log.save()
                    imported_count += 1
        except Exception as e:
            return Response({"error": f"Import failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({"success": True, "imported_count": imported_count})

    @action(detail=False, methods=["get"])
    def export_report(self, request):
        queryset = self.get_queryset().prefetch_related("items__activity_type")
        export_format = request.query_params.get("format", "excel")

        if export_format == "pdf":
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors

            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = 'attachment; filename="seo_work_logs.pdf"'

            # Use landscape layout for wide tables (10 columns)
            doc = SimpleDocTemplate(response, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#1F4E78'),
                spaceAfter=15
            )

            cell_header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Normal'],
                fontSize=8,
                fontName='Helvetica-Bold',
                textColor=colors.whitesmoke,
                alignment=1 # Center
            )

            cell_body_style = ParagraphStyle(
                'BodyStyle',
                parent=styles['Normal'],
                fontSize=7,
                leading=9
            )

            story = []
            story.append(Paragraph("SEO Daily Work Log Report", title_style))
            story.append(Spacer(1, 10))

            headers = ["Date", "Executive", "Website", "Activity Type", "Keyword", "Submission URL", "Username", "DA", "Spam", "Time (m)"]
            table_data = [[Paragraph(h, cell_header_style) for h in headers]]

            for log in queryset:
                for item in log.items.all():
                    table_data.append([
                        Paragraph(log.log_date.strftime("%Y-%m-%d") if log.log_date else "", cell_body_style),
                        Paragraph(log.executive.name or log.executive.username, cell_body_style),
                        Paragraph(log.website.website_name, cell_body_style),
                        Paragraph(item.activity_type.name, cell_body_style),
                        Paragraph(item.keyword or "", cell_body_style),
                        Paragraph(item.submission_url or "", cell_body_style),
                        Paragraph(item.username or "", cell_body_style),
                        Paragraph(str(item.domain_authority) if item.domain_authority is not None else "", cell_body_style),
                        Paragraph(str(item.spam_score) if item.spam_score is not None else "", cell_body_style),
                        Paragraph(str(item.time_spent_minutes) if item.time_spent_minutes is not None else "", cell_body_style)
                    ])

            # Total landscape page width = 792. Margins 30+30=60. Printable width = 732.
            t = Table(table_data, colWidths=[60, 75, 75, 75, 75, 217, 75, 25, 25, 30])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D3D3D3')),
            ]))
            story.append(t)
            doc.build(story)
            return response

        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "SEO Work Logs"

            headers = [
                "Date", "Executive", "Website", "Activity Type", "Keyword", 
                "Submission URL", "Username", "Password", "Domain Authority (DA)", 
                "Spam Score", "Time Spent (Minutes)", "Count"
            ]
            ws.append(headers)

            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            user = request.user
            role_name = user.role.name if getattr(user, 'role', None) else None
            is_admin_user = user.is_superuser or role_name == "SUPER_ADMIN"
            is_seo_manager_user = role_name == "SEO_MANAGER"

            for log in queryset:
                for item in log.items.all():
                    is_submitting = (log.executive == user)
                    decrypted_pw = ""
                    if is_admin_user or is_seo_manager_user or is_submitting:
                        if item.password:
                            decrypted_pw = decrypt_password(item.password)

                    row = [
                        log.log_date.strftime("%Y-%m-%d") if log.log_date else "",
                        log.executive.name or log.executive.username,
                        log.website.website_name,
                        item.activity_type.name,
                        item.keyword or "",
                        item.submission_url or "",
                        item.username or "",
                        decrypted_pw,
                        item.domain_authority if item.domain_authority is not None else "",
                        item.spam_score if item.spam_score is not None else "",
                        item.time_spent_minutes if item.time_spent_minutes is not None else "",
                        item.count if item.count is not None else 1
                    ]
                    ws.append(row)

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="seo_work_logs.xlsx"'
            wb.save(response)
            return response


class SEOMonthlyTargetViewSet(viewsets.ModelViewSet):
    queryset = SEOMonthlyTarget.objects.all()
    serializer_class = SEOMonthlyTargetSerializer

    def get_queryset(self):
        user = self.request.user
        if is_admin(user) or is_seo_manager(user):
            return SEOMonthlyTarget.objects.all().select_related("executive", "website", "activity_type")
        elif is_seo_executive(user):
            return SEOMonthlyTarget.objects.filter(executive=user).select_related("executive", "website", "activity_type")
        return SEOMonthlyTarget.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can create targets.")
        serializer.save()


class SEOTaskPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class SEOTaskViewSet(viewsets.ModelViewSet):
    queryset = SEOTask.objects.all()
    serializer_class = SEOTaskSerializer
    pagination_class = SEOTaskPagination

    def get_queryset(self):
        user = self.request.user
        if is_admin(user) or is_seo_manager(user):
            return SEOTask.objects.all().select_related("website", "assigned_executive", "created_by", "activity_type")
        elif is_seo_executive(user):
            return SEOTask.objects.filter(assigned_executive=user).select_related("website", "assigned_executive", "created_by", "activity_type")
        return SEOTask.objects.none()

    def filter_seo_tasks(self, queryset, exclude_status=False):
        params = self.request.query_params

        website = params.get("website")
        if website:
            queryset = queryset.filter(website_id=website)

        executive = params.get("executive")
        if executive:
            queryset = queryset.filter(assigned_executive_id=executive)

        priority = params.get("priority")
        if priority:
            queryset = queryset.filter(priority=priority)

        activity_type = params.get("activity_type")
        if activity_type:
            queryset = queryset.filter(activity_type_id=activity_type)

        search_query = params.get("search")
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(website__website_name__icontains=search_query) |
                Q(assigned_executive__name__icontains=search_query) |
                Q(assigned_executive__username__icontains=search_query)
            )

        if not exclude_status:
            status_param = params.get("status")
            if status_param:
                if status_param == "overdue":
                    queryset = queryset.filter(due_date__lt=datetime.date.today()).exclude(status="completed")
                else:
                    queryset = queryset.filter(status=status_param)

        return queryset

    def get_ordered_queryset(self, queryset):
        ordering = self.request.query_params.get("ordering")
        ordering_whitelist = {
            "newest": "-created_at",
            "oldest": "created_at",
            "due_date": "due_date",
            "priority": "priority",
            "title": "title"
        }

        db_ordering = "-created_at"
        if ordering in ordering_whitelist:
            db_ordering = ordering_whitelist[ordering]

        if db_ordering == "priority":
            from django.db.models import Case, When, Value, IntegerField
            queryset = queryset.annotate(
                priority_order=Case(
                    When(priority="high", then=Value(1)),
                    When(priority="medium", then=Value(2)),
                    When(priority="low", then=Value(3)),
                    default=Value(4),
                    output_field=IntegerField(),
                )
            ).order_by("priority_order", "-created_at")
        else:
            queryset = queryset.order_by(db_ordering)

        return queryset

    def list(self, request, *args, **kwargs):
        base_qs = self.get_queryset()

        stats_qs = self.filter_seo_tasks(base_qs, exclude_status=True)
        today = datetime.date.today()

        total_count = stats_qs.count()
        pending_count = stats_qs.filter(status="pending").count()
        in_progress_count = stats_qs.filter(status="in_progress").count()
        completed_count = stats_qs.filter(status="completed").count()
        overdue_count = stats_qs.filter(due_date__lt=today).exclude(status="completed").count()

        results_qs = self.filter_seo_tasks(base_qs, exclude_status=False)
        results_qs = self.get_ordered_queryset(results_qs)

        page = self.paginate_queryset(results_qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data["stats"] = {
                "total": total_count,
                "pending": pending_count,
                "in_progress": in_progress_count,
                "completed": completed_count,
                "overdue": overdue_count
            }
            return response

        serializer = self.get_serializer(results_qs, many=True)
        return Response({
            "results": serializer.data,
            "stats": {
                "total": total_count,
                "pending": pending_count,
                "in_progress": in_progress_count,
                "completed": completed_count,
                "overdue": overdue_count
            }
        })

    def perform_create(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can assign SEO tasks.")
        serializer.save(created_by=user)


class SEOReminderViewSet(viewsets.ModelViewSet):
    queryset = SEOReminder.objects.all()
    serializer_class = SEOReminderSerializer

    def get_queryset(self):
        user = self.request.user
        if is_admin(user) or is_seo_manager(user):
            return SEOReminder.objects.all().select_related("website", "assigned_executive", "created_by")
        elif is_seo_executive(user):
            return SEOReminder.objects.filter(assigned_executive=user).select_related("website", "assigned_executive", "created_by")
        return SEOReminder.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can set SEO reminders.")
        serializer.save(created_by=user)


class SEOCredentialViewSet(viewsets.ModelViewSet):
    queryset = SEOCredential.objects.all()
    serializer_class = SEOCredentialSerializer

    def get_queryset(self):
        user = self.request.user
        if not user.is_authenticated:
            return SEOCredential.objects.none()
        if is_admin(user) or is_seo_manager(user):
            return SEOCredential.objects.all().select_related("website")
        elif is_seo_executive(user):
            # View credentials only for websites assigned to them
            return SEOCredential.objects.filter(website__assigned_executive=user).select_related("website")
        return SEOCredential.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can create credentials.")
        serializer.save()

    def perform_update(self, serializer):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can update credentials.")
        serializer.save()

    def perform_destroy(self, instance):
        user = self.request.user
        if not (is_admin(user) or is_seo_manager(user)):
            raise PermissionDenied("Only SEO Managers and Admins can delete credentials.")
        instance.delete()