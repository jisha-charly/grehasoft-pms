from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.throttling import UserRateThrottle
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model
from django.db.models import Prefetch
from django.conf import settings

User = get_user_model()

from .models import UserProfile, WorkSession, ActivityLog, AppActivity
from .serializers import (
    UserProfileSerializer,
    WorkSessionSerializer,
    EmployeeStatusSerializer,
    EmployeeDetailedStatusSerializer,
    HeartbeatRequestSerializer,
    HeartbeatResponseSerializer,
    AppActivitySerializer,
)
from .reports import (
    get_daily_report_data,
    get_weekly_report_data,
    get_monthly_report_data,
    get_employee_analytics_data,
    get_reconciliation_report_data,
    get_session_audit_data,
)
from core.permissions import HasPermission
import datetime
from datetime import timedelta
from .utils import (
    is_tracking_enabled,
    get_or_create_user_profile,
    get_or_create_active_session,
    update_session_ping,
    close_session,
    get_employee_status,
    get_all_employees_status,
    toggle_tracking,
    calculate_daily_working_time,
    format_duration,
)


class HeartbeatThrottle(UserRateThrottle):
    """Rate limit heartbeat API to prevent abuse."""
    scope = 'heartbeat'
    rate = '60/minute'  # 1 per second


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def heartbeat(request):
    """
    Heartbeat API endpoint.
    
    Updates last_ping for user's active session if tracking is enabled.
    Creates new session if none exists.
    Supports app activity duration aggregation to minimize database writes.
    """
    user = request.user
    print("HEARTBEAT REQUEST DATA =", request.data)
    
    # Check if tracking is enabled
    if not is_tracking_enabled(user):
        return Response(
            {
                'success': False,
                'message': 'Tracking is disabled for user',
                'session_id': None,
                'status': 'Offline',
            },
            status=status.HTTP_403_FORBIDDEN
        )
    
    # Validate payload via Serializer
    serializer = HeartbeatRequestSerializer(data=request.data)
    if not serializer.is_valid():
        print("HEARTBEAT SERIALIZER ERRORS =", serializer.errors)
        return Response(
            {
                'success': False,
                'message': 'Validation failed',
                'errors': serializer.errors,
                'session_id': None,
                'status': 'Offline',
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Save telemetry and update active session
        session, created = serializer.save(user=user)
        
        # Get current status after desktop tracker fields have been updated
        current_status = session.get_status()
        
        daily_time = calculate_daily_working_time(user)
        profile = get_or_create_user_profile(user)
        
        return Response(
            {
                'success': True,
                'message': 'Heartbeat recorded' if not created else 'New session created',
                'session_id': session.id,
                'status': current_status,
                'screenshots_enabled': profile.screenshots_enabled,
                'total_work_time': format_duration(daily_time),
                'productive_seconds': session.productive_seconds,
                'idle_seconds': session.idle_seconds,
            },
            status=status.HTTP_200_OK
        )
    except Exception as e:
        import traceback
        print("HEARTBEAT SAVE EXCEPTION =", str(e))
        traceback.print_exc()
        return Response(
            {
                'success': False,
                'message': f'Error: {str(e)}',
                'session_id': None,
                'status': 'Offline',
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout(request):
    """
    Logout endpoint.
    
    Closes the active session for the current user.
    """
    user = request.user
    
    try:
        session = WorkSession.objects.get(
            user=user,
            is_active_session=True
        )
        session = close_session(session)
        
        return Response(
            {
                'success': True,
                'message': 'Session closed',
                'session_id': session.id,
            },
            status=status.HTTP_200_OK
        )
    except WorkSession.DoesNotExist:
        return Response(
            {
                'success': False,
                'message': 'No active session found',
            },
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {
                'success': False,
                'message': f'Error: {str(e)}',
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_status(request, user_id=None):
    """
    Get detailed status for one or all employees.
    
    Query params:
        - user_id: Get status for specific user (optional)
    """
    if user_id:
        # Get specific user status (detailed)
        user = get_object_or_404(User, id=user_id)
        status_data = get_employee_status(user, detailed=True)
        serializer = EmployeeDetailedStatusSerializer(status_data)
        return Response(serializer.data, status=status.HTTP_200_OK)
    else:
        # Get all active users status
        try:
            employees_status = get_all_employees_status()
            serializer = EmployeeStatusSerializer(employees_status, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def toggle_user_tracking(request, user_id):
    """
    Toggle tracking for a specific user.
    
    Body (optional):
        - enabled: true/false (if omitted, toggles current state)
    """
    user = get_object_or_404(User, id=user_id)
    
    # Check permissions: user can only toggle own, admin can toggle any
    if request.user.id != user.id and not request.user.is_staff:
        return Response(
            {'error': 'Permission denied'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        enabled = request.data.get('enabled', None)
        profile = toggle_tracking(user, enable=enabled)
        
        serializer = UserProfileSerializer(profile)
        return Response(
            {
                'success': True,
                'message': f'Tracking {"enabled" if profile.is_tracking_enabled else "disabled"}',
                'data': serializer.data,
            },
            status=status.HTTP_200_OK
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_tracking_status(request):
    """
    Get current user's tracking status.
    """
    user = request.user
    
    try:
        profile = UserProfile.objects.select_related('user').get(user=user)
    except UserProfile.DoesNotExist:
        profile, _ = UserProfile.objects.get_or_create(user=user)
        
    serializer = UserProfileSerializer(profile)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def set_track_enable(request):
    """
    Set tracking enabled/disabled for current user.
    
    Body:
        - enabled: true/false
    """
    user = request.user
    enabled = request.data.get('enabled')
    
    if enabled is None:
        return Response(
            {'error': 'enabled parameter required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        profile = toggle_tracking(user, enable=enabled)
        serializer = UserProfileSerializer(profile)
        return Response(
            {
                'success': True,
                'message': f'Tracking {"enabled" if profile.is_tracking_enabled else "disabled"}',
                'data': serializer.data,
            },
            status=status.HTTP_200_OK
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class UserProfileViewSet(viewsets.ModelViewSet):
    """ViewSet for UserProfile management."""
    queryset = UserProfile.objects.select_related('user')
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter by current user unless admin."""
        if self.request.user.is_staff:
            return UserProfile.objects.select_related('user')
        return UserProfile.objects.filter(user=self.request.user)


class WorkSessionViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for WorkSession viewing."""
    queryset = WorkSession.objects.select_related('user')
    serializer_class = WorkSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filter by current user unless admin."""
        user_id = self.request.query_params.get('user_id')
        
        if self.request.user.is_staff:
            if user_id:
                return WorkSession.objects.filter(user_id=user_id).select_related('user')
            return WorkSession.objects.select_related('user')
        
        return WorkSession.objects.filter(user=self.request.user).select_related('user')

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get active session for user."""
        try:
            session = WorkSession.objects.get(
                user=request.user,
                is_active_session=True
            )
            serializer = self.get_serializer(session)
            return Response(serializer.data)
        except WorkSession.DoesNotExist:
            return Response(
                {'message': 'No active session'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['get'])
    def today(self, request):
        """Get all sessions for today."""
        today = timezone.now().date()
        sessions = WorkSession.objects.filter(
            user=request.user,
            login_time__date=today
        ).select_related('user')
        
        serializer = self.get_serializer(sessions, many=True)
        return Response(serializer.data)


from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import serializers as drf_serializers

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def activity_batch_sync(request):
    """
    Sync queued offline activities with duration aggregation.
    """
    user = request.user
    print("Incoming activity batch:", request.data)
    if not is_tracking_enabled(user):
        return Response({'success': False, 'message': 'Tracking is disabled'}, status=status.HTTP_403_FORBIDDEN)
        
    device_id = request.data.get('device_id') or 'default'
    installation_uuid = request.data.get('installation_uuid')
    tracker_id = request.data.get('tracker_id')
    machine_fingerprint = request.data.get('machine_fingerprint')

    session, created = get_or_create_active_session(user, device_id=device_id)
    if created or not session.installation_uuid:
        session.installation_uuid = installation_uuid
        session.tracker_id = tracker_id
        session.machine_fingerprint = machine_fingerprint
        session.save(update_fields=['installation_uuid', 'tracker_id', 'machine_fingerprint'])

    print("SESSION ID:", session.id)
    activities = request.data.get('activities', [])
    
    # Sort activities by timestamp to aggregate in order
    activities = sorted(activities, key=lambda x: x.get('timestamp', ''))
    
    # Align session login_time and last_ping with batch activity timestamps to ensure correct session span
    batch_timestamps = []
    for act in activities:
        ts_str = act.get('timestamp')
        if ts_str:
            try:
                ts = datetime.datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
                batch_timestamps.append((ts, act.get('duration_seconds', 10)))
            except Exception:
                pass
                
    if batch_timestamps:
        min_ts_info = min(batch_timestamps, key=lambda x: x[0])
        max_ts_info = max(batch_timestamps, key=lambda x: x[0])
        min_ts = min_ts_info[0]
        max_ts = max_ts_info[0]
        first_act_duration = min_ts_info[1]
        
        session_login_candidate = min_ts - timedelta(seconds=first_act_duration)
        session_ping_candidate = max_ts
        
        if created:
            session.login_time = session_login_candidate
            session.last_ping = session_ping_candidate
            session.save(update_fields=['login_time', 'last_ping'])
        else:
            session.refresh_from_db()
            if session_login_candidate < session.login_time:
                session.login_time = session_login_candidate
            if session_ping_candidate > session.last_ping:
                session.last_ping = session_ping_candidate
            session.save(update_fields=['login_time', 'last_ping'])
    
    MIN_MOUSE_ACTIVITY = getattr(settings, 'TRACKING_MIN_MOUSE_ACTIVITY', 5)
    MIN_KEY_ACTIVITY = getattr(settings, 'TRACKING_MIN_KEY_ACTIVITY', 1)
    
    synced_count = 0
    total_mouse_moves = 0
    total_key_presses = 0
    total_clicks = 0
    total_productive_seconds = 0
    total_idle_seconds = 0
    
    for act in activities:
        app_name = act.get('app_name')
        window_title = act.get('window_title', '')
        duration_seconds = act.get('duration_seconds', 10)
        is_idle = act.get('is_idle', False)
        mouse_moves = act.get('mouse_moves', 0)
        key_presses = act.get('key_presses', 0)
        clicks = act.get('clicks', 0)
        timestamp_str = act.get('timestamp')
        
        if not app_name:
            continue
            
        timestamp = timezone.now()
        if timestamp_str:
            try:
                timestamp = datetime.datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            except Exception:
                pass

        # Sanitize/Convert types to ensure robustness
        try:
            duration_seconds = int(duration_seconds)
        except (ValueError, TypeError):
            duration_seconds = 10
        try:
            mouse_moves = int(mouse_moves)
        except (ValueError, TypeError):
            mouse_moves = 0
        try:
            key_presses = int(key_presses)
        except (ValueError, TypeError):
            key_presses = 0
        try:
            clicks = int(clicks)
        except (ValueError, TypeError):
            clicks = 0
        is_idle = bool(is_idle)
        
        # Log task items
        print("Batch Item Details:")
        print("is_idle:", is_idle)
        print("mouse_moves:", mouse_moves)
        print("key_presses:", key_presses)
        print("clicks:", clicks)
        
        from .reports import classify_app_activity
        category = classify_app_activity(app_name, window_title)
        is_productive_app = (category != 'non_productive')
        print("app_name:", app_name)
        print("is_productive_app:", is_productive_app)
        
        is_productive_tick = (not is_idle) and (mouse_moves > 0 or key_presses > 0 or clicks > 0)
        if is_productive_tick:
            print("ENTERED PRODUCTIVE BRANCH")
            print("ADDING:", duration_seconds)
            tick_productive_seconds = duration_seconds
            tick_idle_seconds = 0
        else:
            print("ENTERED IDLE BRANCH")
            print("ADDING IDLE:", duration_seconds)
            tick_productive_seconds = 0
            tick_idle_seconds = duration_seconds
        
        total_mouse_moves += mouse_moves
        total_key_presses += key_presses
        total_clicks += clicks
        total_productive_seconds += tick_productive_seconds
        total_idle_seconds += tick_idle_seconds
        
        # Aggregate duration if active app/window hasn't changed
        last_activity = AppActivity.objects.filter(session=session).order_by('-timestamp').first()
        if last_activity and last_activity.app_name == app_name and last_activity.window_title == window_title:
            last_activity.duration_seconds += duration_seconds
            last_activity.mouse_moves += mouse_moves
            last_activity.key_presses += key_presses
            last_activity.clicks += clicks
            last_activity.productive_seconds += tick_productive_seconds
            last_activity.productive_duration += tick_productive_seconds
            last_activity.timestamp = timestamp
            last_activity.save(update_fields=['duration_seconds', 'mouse_moves', 'key_presses', 'clicks', 'productive_seconds', 'productive_duration', 'timestamp'])
        else:
            AppActivity.objects.create(
                user=user,
                session=session,
                app_name=app_name,
                window_title=window_title,
                duration_seconds=duration_seconds,
                mouse_moves=mouse_moves,
                key_presses=key_presses,
                clicks=clicks,
                productive_seconds=tick_productive_seconds,
                productive_duration=tick_productive_seconds,
                timestamp=timestamp
            )
        synced_count += 1
        
    # Refresh session from database before incrementing fields to prevent race conditions
    session.refresh_from_db()
    
    session.mouse_moves += total_mouse_moves
    session.key_presses += total_key_presses
    session.clicks += total_clicks
    session.productive_seconds += total_productive_seconds
    session.idle_seconds += total_idle_seconds
    session.tracked_seconds = session.productive_seconds + session.idle_seconds
    
    if session.tracked_seconds > 0:
        session.activity_percentage = min(100.0, (session.productive_seconds / session.tracked_seconds) * 100.0)
    else:
        session.activity_percentage = 0.0
        
    from .reports import detect_breaks_and_gaps
    today = timezone.now().date()
    break_analysis = detect_breaks_and_gaps(user, today, today, sessions_list=[session])
    session.break_count = break_analysis['break_count']
    
    session.save(update_fields=[
        'mouse_moves', 'key_presses', 'clicks', 'productive_seconds', 
        'idle_seconds', 'tracked_seconds', 'break_count', 'activity_percentage', 'updated_at'
    ])
    
    session.refresh_from_db()
    print("Database Save (Batch Sync) - session.productive_seconds:", session.productive_seconds)
    print("Database Save (Batch Sync) - session.idle_seconds:", session.idle_seconds)
    print("Database Save (Batch Sync) - session.activity_percentage:", session.activity_percentage)
    print("DB productive_seconds:", session.productive_seconds)
    print("productive_seconds:", session.productive_seconds)
    print("idle_seconds:", session.idle_seconds)
    
    update_session_ping(session)
    daily_time = calculate_daily_working_time(user)
        
    return Response({
        'success': True,
        'synced_count': synced_count,
        'total_work_time': format_duration(daily_time),
        'productive_seconds': session.productive_seconds,
        'idle_seconds': session.idle_seconds,
    }, status=status.HTTP_200_OK)



import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from rest_framework.negotiation import DefaultContentNegotiation

class IgnoreFormatContentNegotiation(DefaultContentNegotiation):
    def select_renderer(self, request, renderers, format_suffix=None):
        # Always return the default renderer to ignore 'format' query parameters for suffix negotiation
        return (renderers[0], renderers[0].media_type)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
def daily_report_view(request):
    """Get daily reports for employees."""
    # Set permission check
    request.parser_context['view'].required_permission = 'MANAGE_SETTINGS'
    
    date_str = request.query_params.get('date')
    if date_str:
        try:
            start_date = datetime.date.fromisoformat(date_str)
            end_date = start_date
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Default to today
        start_date = timezone.now().date()
        end_date = start_date
        
    start_date_param = request.query_params.get('start_date')
    end_date_param = request.query_params.get('end_date')
    if start_date_param and end_date_param:
        try:
            start_date = datetime.date.fromisoformat(start_date_param)
            end_date = datetime.date.fromisoformat(end_date_param)
        except ValueError:
            return Response({'error': 'Invalid start_date or end_date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

    dept_id = request.query_params.get('department_id')
    search = request.query_params.get('search')
    
    data = get_daily_report_data(start_date, end_date, department_id=dept_id, search_query=search)
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
def reconciliation_report_view(request):
    """Get reconciliation report for employees."""
    request.parser_context['view'].required_permission = 'MANAGE_SETTINGS'
    
    start_date_param = request.query_params.get('start_date')
    end_date_param = request.query_params.get('end_date')
    
    if start_date_param and end_date_param:
        try:
            start_date = datetime.date.fromisoformat(start_date_param)
            end_date = datetime.date.fromisoformat(end_date_param)
        except ValueError:
            return Response({'error': 'Invalid start_date or end_date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Default to today
        start_date = timezone.now().date()
        end_date = start_date

    dept_id = request.query_params.get('department_id')
    search = request.query_params.get('search')
    
    data = get_reconciliation_report_data(start_date, end_date, department_id=dept_id, search_query=search)
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
def session_audit_report_view(request):
    """Get historical session audit report."""
    request.parser_context['view'].required_permission = 'MANAGE_SETTINGS'
    
    start_date_param = request.query_params.get('start_date')
    end_date_param = request.query_params.get('end_date')
    
    if start_date_param and end_date_param:
        try:
            start_date = datetime.date.fromisoformat(start_date_param)
            end_date = datetime.date.fromisoformat(end_date_param)
        except ValueError:
            return Response({'error': 'Invalid start_date or end_date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Default to today
        start_date = timezone.now().date()
        end_date = start_date

    user_id = request.query_params.get('user_id')
    if user_id:
        try:
            user_id = int(user_id)
        except ValueError:
            return Response({'error': 'Invalid user_id.'}, status=status.HTTP_400_BAD_REQUEST)

    data = get_session_audit_data(start_date, end_date, user_id=user_id)
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
def weekly_report_view(request):
    """Get weekly summaries and trends."""
    request.parser_context['view'].required_permission = 'MANAGE_SETTINGS'
    
    start_date_str = request.query_params.get('start_date')
    if start_date_str:
        try:
            start_date = datetime.date.fromisoformat(start_date_str)
        except ValueError:
            return Response({'error': 'Invalid start_date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Default to 7 days ago
        start_date = timezone.now().date() - timedelta(days=6)
        
    end_date_str = request.query_params.get('end_date')
    if end_date_str:
        try:
            end_date = datetime.date.fromisoformat(end_date_str)
        except ValueError:
            return Response({'error': 'Invalid end_date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        end_date = start_date + timedelta(days=6)

    dept_id = request.query_params.get('department_id')
    
    data = get_weekly_report_data(start_date, end_date, department_id=dept_id)
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
def monthly_report_view(request):
    """Get monthly summaries and rankings."""
    request.parser_context['view'].required_permission = 'MANAGE_SETTINGS'
    
    year_str = request.query_params.get('year')
    month_str = request.query_params.get('month')
    
    now = timezone.now()
    year = int(year_str) if year_str else now.year
    month = int(month_str) if month_str else now.month
    
    if month < 1 or month > 12:
        return Response({'error': 'Month must be between 1 and 12.'}, status=status.HTTP_400_BAD_REQUEST)
        
    dept_id = request.query_params.get('department_id')
    
    data = get_monthly_report_data(year, month, department_id=dept_id)
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
def employee_analytics_view(request):
    """Get detailed report for a specific employee."""
    request.parser_context['view'].required_permission = 'MANAGE_SETTINGS'
    
    user_id = request.query_params.get('user_id')
    if not user_id:
        return Response({'error': 'user_id query parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)
        
    user = get_object_or_404(User, id=user_id)
    
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.date.fromisoformat(start_date_str)
            end_date = datetime.date.fromisoformat(end_date_str)
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Default to last 7 days
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=6)
        
    data = get_employee_analytics_data(user, start_date, end_date)
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, HasPermission])
def export_report_view(request):
    """Export work tracking reports to CSV, Excel, or PDF."""
    request.parser_context['view'].required_permission = 'MANAGE_SETTINGS'
    
    export_format = request.query_params.get('format', 'csv').lower()
    report_type = request.query_params.get('type', 'daily').lower()
    
    # Common filters
    date_str = request.query_params.get('date')
    start_date_str = request.query_params.get('start_date')
    end_date_str = request.query_params.get('end_date')
    dept_id = request.query_params.get('department_id')
    search = request.query_params.get('search')
    user_id = request.query_params.get('user_id')
    
    # Calculate dates
    today = timezone.now().date()
    start_date = today
    end_date = today
    
    if date_str:
        try:
            start_date = datetime.date.fromisoformat(date_str)
            end_date = start_date
        except ValueError:
            pass
    elif start_date_str and end_date_str:
        try:
            start_date = datetime.date.fromisoformat(start_date_str)
            end_date = datetime.date.fromisoformat(end_date_str)
        except ValueError:
            pass
            
    # Gather Data
    if report_type == 'daily':
        data = get_daily_report_data(start_date, end_date, department_id=dept_id, search_query=search)
        filename = f"daily_report_{start_date}"
    elif report_type == 'reconciliation':
        data = get_reconciliation_report_data(start_date, end_date, department_id=dept_id, search_query=search)
        filename = f"reconciliation_report_{start_date}"
    elif report_type == 'weekly':
        data_summary = get_weekly_report_data(start_date, end_date, department_id=dept_id)
        # Flatten daily productivity trend for weekly rows
        data = []
        for trend in data_summary['daily_productivity_trend']:
            data.append({
                'Date': trend['date'],
                'Productive Hours': trend['productive_hours'],
                'Idle Hours': trend['idle_hours'],
                'Total Tracked': round(trend['productive_hours'] + trend['idle_hours'], 2)
            })
        filename = f"weekly_report_{start_date}_to_{end_date}"
    elif report_type == 'monthly':
        year = int(request.query_params.get('year', today.year))
        month = int(request.query_params.get('month', today.month))
        data_summary = get_monthly_report_data(year, month, department_id=dept_id)
        data = data_summary['employee_ranking']
        filename = f"monthly_ranking_{year}_{month}"
    elif report_type == 'employee' and user_id:
        user = get_object_or_404(User, id=user_id)
        data_summary = get_employee_analytics_data(user, start_date, end_date)
        data = data_summary['daily_breakdown']
        filename = f"employee_report_{user.username}_{start_date}_to_{end_date}"
    else:
        return Response({'error': 'Invalid report type or missing parameters.'}, status=status.HTTP_400_BAD_REQUEST)

    # 1. EXPORT CSV
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        if not data:
            return response
            
        writer = csv.writer(response)
        if report_type in ['daily', 'reconciliation', 'employee']:
            headers = [
                'Employee Name', 'Employee Code', 'Department', 'Date',
                'Productive Time', 'Idle Time', 'Desktop Work Time', 'Portal Active Time',
                'Break Time', 'Unaccounted Time', 'Total Engagement Time', 'Workday Span',
                'Activity Percentage', 'Status'
            ]
            keys = [
                'employee_name', 'employee_code', 'department', 'date',
                'productive_time', 'idle_time', 'desktop_work_time', 'portal_active_time',
                'break_time', 'unaccounted_time', 'total_engagement_time', 'workday_span',
                'activity_percentage', 'status'
            ]
            writer.writerow(headers)
            for row in data:
                writer.writerow([row.get(k, '-') for k in keys])
        else:
            headers = list(data[0].keys())
            writer.writerow(headers)
            for row in data:
                writer.writerow([row[h] for h in headers])
            
        return response

    # 2. EXPORT EXCEL
    elif export_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.title()
        
        if not data:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            wb.save(response)
            return response
            
        if report_type in ['daily', 'reconciliation', 'employee']:
            headers = [
                'Employee Name', 'Employee Code', 'Department', 'Date',
                'Productive Time', 'Idle Time', 'Desktop Work Time', 'Portal Active Time',
                'Break Time', 'Unaccounted Time', 'Total Engagement Time', 'Workday Span',
                'Activity Percentage', 'Status'
            ]
            keys = [
                'employee_name', 'employee_code', 'department', 'date',
                'productive_time', 'idle_time', 'desktop_work_time', 'portal_active_time',
                'break_time', 'unaccounted_time', 'total_engagement_time', 'workday_span',
                'activity_percentage', 'status'
            ]
        else:
            headers = [h.replace('_', ' ').title() for h in data[0].keys() if not h.startswith('raw_')]
            keys = [k for k in data[0].keys() if not k.startswith('raw_')]
            
        ws.append(headers)
        
        # Style Header
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Add Data Rows
        for item in data:
            row_data = [item.get(k, '-') for k in keys]
            ws.append(row_data)
            
        # Adjust Columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # 3. EXPORT PDF
    elif export_format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        doc = SimpleDocTemplate(
            response, 
            pagesize=landscape(A4),
            rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=15
        )
        
        story = []
        
        title_p = Paragraph(f"Grehasoft Work Tracking - {report_type.title()} Report ({start_date} to {end_date})", title_style)
        story.append(title_p)
        story.append(Spacer(1, 10))
        
        if not data:
            story.append(Paragraph("No tracking data recorded for this period.", styles['Normal']))
            doc.build(story)
            return response
            
        # Define customized column layout for each report type. Use newlines (\n) for headers in 14-column reports.
        report_columns = {
            'daily': [
                ('employee_name', 'Employee\nName'),
                ('employee_code', 'Employee\nCode'),
                ('department', 'Department'),
                ('date', 'Date'),
                ('productive_time', 'Productive\nTime'),
                ('idle_time', 'Idle\nTime'),
                ('desktop_work_time', 'Desktop\nWork'),
                ('portal_active_time', 'Portal\nActive'),
                ('break_time', 'Break\nTime'),
                ('unaccounted_time', 'Unaccounted\nTime'),
                ('total_engagement_time', 'Total\nEngagement'),
                ('workday_span', 'Workday\nSpan'),
                ('activity_percentage', 'Activity\n%'),
                ('status', 'Status')
            ],
            'reconciliation': [
                ('employee_name', 'Employee\nName'),
                ('employee_code', 'Employee\nCode'),
                ('department', 'Department'),
                ('date', 'Date'),
                ('productive_time', 'Productive\nTime'),
                ('idle_time', 'Idle\nTime'),
                ('desktop_work_time', 'Desktop\nWork'),
                ('portal_active_time', 'Portal\nActive'),
                ('break_time', 'Break\nTime'),
                ('unaccounted_time', 'Unaccounted\nTime'),
                ('total_engagement_time', 'Total\nEngagement'),
                ('workday_span', 'Workday\nSpan'),
                ('activity_percentage', 'Activity\n%'),
                ('status', 'Status')
            ],
            'weekly': [
                ('Date', 'Date'),
                ('Productive Hours', 'Productive\nHours'),
                ('Idle Hours', 'Idle\nHours'),
                ('Total Tracked', 'Total\nTracked')
            ],
            'monthly': [
                ('full_name', 'Full Name'),
                ('employee_code', 'Employee Code'),
                ('department', 'Department'),
                ('productive_hours', 'Productive Hours'),
                ('tracked_hours', 'Tracked Hours'),
                ('activity_percentage', 'Activity %')
            ],
            'employee': [
                ('employee_name', 'Employee\nName'),
                ('employee_code', 'Employee\nCode'),
                ('department', 'Department'),
                ('date', 'Date'),
                ('productive_time', 'Productive\nTime'),
                ('idle_time', 'Idle\nTime'),
                ('desktop_work_time', 'Desktop\nWork'),
                ('portal_active_time', 'Portal\nActive'),
                ('break_time', 'Break\nTime'),
                ('unaccounted_time', 'Unaccounted\nTime'),
                ('total_engagement_time', 'Total\nEngagement'),
                ('workday_span', 'Workday\nSpan'),
                ('activity_percentage', 'Activity\n%'),
                ('status', 'Status')
            ]
        }
        
        cols = report_columns.get(report_type)
        if cols:
            keys = [c[0] for c in cols if c[0] in data[0]]
            headers = [c[1] for c in cols if c[0] in data[0]]
        else:
            keys = [k for k in data[0].keys() if not k.startswith('raw_')][:8]
            headers = [k.replace('_', ' ').title() for k in keys]
            
        col_widths = None
        is_large_report = report_type in ['daily', 'reconciliation', 'employee'] and len(keys) == 14
        
        if is_large_report:
            # A4 landscape width is 841.89 points.
            # Printable width with left/right margins of 30 is 781.89 points.
            # Explicit column widths totaling 749 points:
            col_widths = [90, 48, 80, 55, 48, 48, 48, 48, 48, 48, 48, 48, 44, 48]

        # Adjust font size and padding if we have many columns
        font_size_header = 7 if is_large_report else 10
        font_size_cell = 6 if is_large_report else 8
        padding_val = 3 if is_large_report else 6

        # Create paragraph style for table headers to support multi-line text wrapping (ReportLab Table cells need Flowables for newlines/br)
        header_text_style = ParagraphStyle(
            'HeaderTextStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=font_size_header,
            textColor=colors.whitesmoke,
            alignment=1,  # Center alignment
            leading=font_size_header + 2  # Adjust line height dynamically
        )
        
        # Wrap header labels in Paragraph to render HTML <br/> line breaks
        wrapped_headers = [Paragraph(h.replace('\n', '<br/>'), header_text_style) for h in headers]
        
        table_data = [wrapped_headers]
        for item in data:
            row = [str(item.get(k) if item.get(k) is not None else '-') for k in keys]
            table_data.append(row)
            
        t = Table(table_data, colWidths=col_widths)
        
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), font_size_header),
            ('TOPPADDING', (0,0), (-1,-1), padding_val),
            ('BOTTOMPADDING', (0,0), (-1,-1), padding_val),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#F8FAFC')),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F1F5F9')]),
            ('FONTSIZE', (0,1), (-1,-1), font_size_cell),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        
        story.append(t)
        doc.build(story)
        return response
        
    return Response({'error': 'Invalid format requested.'}, status=status.HTTP_400_BAD_REQUEST)


# Bind custom content negotiation class to export view
export_report_view.cls.content_negotiation_class = IgnoreFormatContentNegotiation

# Secure report views by setting class-level required_permission
daily_report_view.cls.required_permission = 'MANAGE_SETTINGS'
weekly_report_view.cls.required_permission = 'MANAGE_SETTINGS'
monthly_report_view.cls.required_permission = 'MANAGE_SETTINGS'
employee_analytics_view.cls.required_permission = 'MANAGE_SETTINGS'
export_report_view.cls.required_permission = 'MANAGE_SETTINGS'
reconciliation_report_view.cls.required_permission = 'MANAGE_SETTINGS'
session_audit_report_view.cls.required_permission = 'MANAGE_SETTINGS'

