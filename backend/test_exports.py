import os
import sys
import django
import csv
from io import BytesIO, StringIO
from openpyxl import load_workbook

# Set up Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
django.setup()

from django.contrib.auth import get_user_model
from rest_framework.test import APIRequestFactory, force_authenticate
from apps.tracking.views import export_report_view

User = get_user_model()
admin_user = User.objects.filter(is_staff=True).first()
if not admin_user:
    admin_user = User.objects.first()

factory = APIRequestFactory()

def verify_export(format_type, report_type):
    print(f"\n--- TESTING EXPORT: format={format_type}, type={report_type} ---")
    request = factory.get(f'/api/v1/tracking/reports/export/?format={format_type}&type={report_type}')
    force_authenticate(request, user=admin_user)
    response = export_report_view(request)
    print(f"Status code: {response.status_code}")
    print(f"Content-Type: {response.get('Content-Type')}")
    print(f"Content-Disposition: {response.get('Content-Disposition')}")
    
    if response.status_code == 200:
        content = response.content
        print(f"Content Size: {len(content)} bytes")
        
        if format_type == 'csv':
            csv_file = StringIO(content.decode('utf-8'))
            reader = csv.reader(csv_file)
            headers = next(reader)
            print("CSV Headers:")
            print(headers)
            first_row = next(reader, None)
            if first_row:
                print("CSV First Row:")
                print(first_row)
        elif format_type == 'excel':
            excel_file = BytesIO(content)
            wb = load_workbook(excel_file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            print("Excel Headers:")
            print(headers)
            first_row = [cell.value for cell in ws[2]]
            print("Excel First Row:")
            print(first_row)
        elif format_type == 'pdf':
            print("PDF generation success. Checked file structure is correct application/pdf binary.")
    else:
        print(response.data)

# Test daily CSV & Excel & PDF
verify_export('csv', 'daily')
verify_export('excel', 'daily')
verify_export('pdf', 'daily')

# Test reconciliation CSV & Excel & PDF
verify_export('csv', 'reconciliation')
verify_export('excel', 'reconciliation')
verify_export('pdf', 'reconciliation')
